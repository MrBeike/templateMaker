from ctypes import alignment
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
from openpyxl.utils import get_column_letter

fullPath = input("Enter the name of the file: ")
filePath,fileName = os.path.split(fullPath)
f = pd.ExcelFile(fullPath)
sheet = f.sheet_names
newFile = os.path.join(filePath,f"new{fileName}")
writer = pd.ExcelWriter(newFile)

dataSet = []
for i in sheet:
    data = pd.read_excel(fullPath, sheet_name=i)
    # 解除单元格合并，前置填充
    data["字段中文名称"] = data["字段中文名称"].ffill()
    # 利用groupby合并同样分组，并汇总填写规则（合并单元格实质还是多个单元格，此处合并为一个单元格，sort=False保持原来顺序）
    dataNew = data.groupby(["字段中文名称"],sort=False)['填写规则'].apply(list).to_frame()
    # 前面利用list合并填写规则，此处将list转换为字符串【lambda x: ''.join(x)）报错，暂时没弄明白】
    dataNew['填写规则'] = dataNew['填写规则'].apply(lambda x:str(x).replace("[","").replace("]","").replace("'",""))
    # 转置
    dataTransport = dataNew.T
    # 写入文件，注意index=False，不写入索引
    dataTransport.to_excel(writer, sheet_name=i,index=False)
writer.save()
writer.close()

f = pd.ExcelFile(newFile)
sheet = f.sheet_names
wb = load_workbook(newFile)
# 将第二行的数据当成批注写入标题行
for i in sheet:
    ws = wb[i]
    for column in range(1,ws.max_column+1):
        # 获取第二行单元格内容
        commentContent = ws.cell(row=2,column=column).value
        # 构建Comment对象,（内容，作者）
        comment = Comment(commentContent, 'xeroxYor')
        # 添加批注
        ws.cell(row=1,column=column).comment = comment
    # 删除第二行内容   
    ws.delete_rows(2)
    # 恢复第二行行高到Excel默认值
    ws.row_dimensions[2].height = 13.5
    # 设置第一行的背景色、列宽、行高等 I love this blue :tada:
    fill = PatternFill("solid", fgColor="FFD0E9F5")
    side = Side(border_style="thin", color="FF3BA5D3")
    border = Border(left=side,right=side,top=side,bottom=side)
    alignment = Alignment(horizontal="center",vertical="center",wrapText=True)
    for i in range(1,ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].width = 20
        ws.cell(row=1,column=i).fill = fill
        ws.cell(row=1,column=i).alignment = alignment
        ws.cell(row=1,column=i).border = border
    ws.row_dimensions[1].height = 22.5
    # 设置第一行的单元格字体
    ws.row_dimensions[1].font = Font(size=12,bold=True)
wb.save(os.path.join(filePath,f"【template】{fileName}"))

